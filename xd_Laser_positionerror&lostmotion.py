from openpyxl import load_workbook
from openpyxl.styles import Alignment


# 1부터 n행까지 삭제
def del_row(num_row=18):
    for i in range(num_row, 0, -1):  # 18행부터 1행까지 역순으로 삭제
        sheet.delete_rows(i)

    print("불필요한 행을 제거했습니다.")


# m열 이상 삭제
def del_col(num_col=4):
    # 'D'열의 인덱스는 4 (A=1, B=2, C=3, D=4)
    start_delete_column_index = num_col

    # 마지막 열의 인덱스
    end_delete_column_index = sheet.max_column

    # 'D'열부터 마지막 열까지 삭제
    sheet.delete_cols(
        start_delete_column_index,
        end_delete_column_index - start_delete_column_index + 1,
    )

    print("열 D부터 마지막 열까지 삭제되었습니다.")


# 특정 단어를 찾아서 바꾸기(열 제목 바꾸기)
def replace_text_in_cells(file, sheet, replacements):
    # 엑셀 파일 열기
    # workbook = openpyxl.load_workbook(file)

    # 원하는 시트 선택
    # sheet = workbook[sheet]

    # 각 셀을 순회하며 특정 문자열 또는 숫자를 찾아 대체하는 작업 수행
    for row in sheet.iter_rows():
        for cell in row:
            cell_value = str(cell.value)  # 셀 값 가져오기 (문자열로 변환)
            for find_value, replace_value in replacements.items():
                if str(find_value) in cell_value:
                    cell.value = None  # cell 지운다
                    # cell_value = cell_value.replace(
                    # str(find_value), str(replace_value)
                    # )
                    cell.value = replace_value

    print("셀 내용이 변경되었습니다.")


"""
프로그램 실행
"""
# 엑셀 파일 선택
file_path = "C:/Users/goodk/Documents/Automatiom Test with ChatGPT.xlsx"
wb = load_workbook(file_path)  # 엑셀 파일 로드
sheet = wb["Law Data"]  # 시트 선택

del_row(num_row=18)  # 필요없는 행 삭제
replacements = {  # 엑셀 파일 경로, 시트 이름, 찾을 값 및 대체 값으로 이루어진 딕셔너리 지정
    "Position": "지령위치",
    "LasRead": "실제위치",
    "LinearError": "오차",
    12345: 99999,  # 숫자도 가능합니다.
    "Status": 1234,
}
replace_text_in_cells(file_path, sheet, replacements)  # 찾아 바꾸기(제목행 한글로 변경)


# 변경 사항 저장
wb.save(file_path)
wb.close()
