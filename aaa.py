import openpyxl


def replace_text_in_cells(workbook_path, sheet_name, replacements):
    # 엑셀 파일 열기
    workbook = openpyxl.load_workbook(workbook_path)

    # 원하는 시트 선택
    sheet = workbook[sheet_name]

    # 각 셀을 순회하며 특정 문자열 또는 숫자를 찾아 삭제하고 대체하는 작업 수행
    for row in sheet.iter_rows():
        for cell in row:
            cell_value = str(cell.value)  # 셀 값 가져오기 (문자열로 변환)
            for find_value, replace_value in replacements.items():
                if str(find_value) in cell_value:
                    cell.value = None  # 셀 값을 삭제합니다.

    # 엑셀 파일 저장
    workbook.save(workbook_path)

    # 삭제된 값을 대체할 데이터로 입력
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                cell.value = replacements.get(cell.value, "")  # 대체할 데이터로 채웁니다.

    # 엑셀 파일 다시 저장
    workbook.save(workbook_path)

    # 엑셀 파일 닫기
    workbook.close()


# 엑셀 파일 경로, 시트 이름, 찾을 값 및 대체 값으로 이루어진 딕셔너리 지정
workbook_path = "C:/Users/goodk/Documents/Automatiom Test with ChatGPT.xlsx"
sheet_name = "Law Data"
replacements = {
    "Position": "지령위치",
    "LasRead": "실제위치",
    "LinearError": "오차",
    12345: 99999,  # 숫자도 가능합니다.
}

# 함수 호출하여 찾을 값과 대체 값으로 대체 작업 수행
replace_text_in_cells(workbook_path, sheet_name, replacements)

print("셀 내용이 변경되었습니다.")
