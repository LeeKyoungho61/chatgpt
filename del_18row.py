from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 엑셀 파일 선택
file = "C:/Users/goodk/Documents/Automatiom Test with ChatGPT.xlsx"

# 엑셀 파일 로드
workbook = load_workbook(file)

# 시트 선택
sheet = workbook["Law Data"]


# 필요없는 열 삭제 #################################################
# 'D'열부터 데이터가 있는 마지막 열까지 삭제하는 작업을 수행.

# 'D'열의 인덱스는 4 (A=1, B=2, C=3, D=4)
start_delete_column_index = 4

# 마지막 열의 인덱스
end_delete_column_index = sheet.max_column

# 'D'열부터 마지막 열까지 삭제
sheet.delete_cols(
    start_delete_column_index,
    end_delete_column_index - start_delete_column_index + 1,
)

print("열 D부터 마지막 열까지 삭제되었습니다.")
# workbook.save(file)

# # 2행(제목행) 가로맞춤을 가운데 정렬로 설정하고,
# # 모든 셀의 세로맞춤을 가운데 정렬로 설정.

# # 2행의 가로맞춤 설정
# for cell in sheet[2]:
#     cell.alignment = Alignment(horizontal="center")

# # 전체 셀의 세로맞춤 설정
# for row in sheet.iter_rows():
#     for cell in row:
#         cell.alignment = Alignment(vertical="center")

# print("제목줄이 정렬되었습니다.")


# """
# A열의 데이터는 반복되고 있다. 0부터 차츰 커지다가 정점에 이르면 다시 작아져서 0이되는 패턴이 반복된다.
# 이렇게 반복되는 데이터를 찾아서 새로운 시트로 복사하자.
# 예를 들면 현재 시트는 A3부터 A128, A129부터 A254, A255부터 A380이 반복되고 있다.
# 3행부터 128행의 데이터를 복사해서 '첫번째 측정데이터' 시트에 저장,
# 129행부터 254행의 데이터를 복사해서 '두번째 측정데이터' 시트에 저장,
# 255행부터 380행의 데이터를 복사해서 '세번째 측정데이터' 시트에 저장해줘.
# 새 시트의 1행과 2행은 원본과 동일해야 해. (1행은 비어있고, 2행은 제목줄이야)
# 이때 반복되는 데이터를 찾는 것이 중요해. 이것이 잘 처리되도록 코딩해줘.

# [챗GPT]
# 반복되는 데이터 패턴을 찾아 새로운 시트에 복사하는 작업을 완료했습니다. 찾아낸 패턴은 다음과 같습니다:

# 첫 번째 패턴: 0행부터 125행까지 (엑셀의 3행부터 128행에 해당)
# 두 번째 패턴: 125행부터 251행까지 (엑셀의 129행부터 254행에 해당)
# 세 번째 패턴: 251행부터 377행까지 (엑셀의 255행부터 380행에 해당)
# 각 패턴의 데이터는 각각 '첫번째 측정데이터', '두번째 측정데이터', '세번째 측정데이터'라는 이름의 새 시트에 저장되었습니다. 원본 시트의 1행과 2행(제목줄)은 새 시트에 동일하게 복사되었습니다.
# """

# # A열의 데이터가 반복되는 패턴을 찾아 각 패턴별로 새 시트에 데이터를 복사하는 작업을 수행합니다.

# # # 엑셀 파일 로드
# # workbook = load_workbook(file)
# # sheet = workbook["Law Data"]

# # A열의 데이터를 가져와 패턴을 찾습니다.
# a_column_data = [
#     cell.value for cell in sheet["A"][header_row:] if cell.value is not None
# ]  # 헤더 다음 행부터 데이터 추출


# # 반복되는 패턴의 시작점과 끝점 인덱스를 찾습니다.
# def find_repeating_patterns(data):
#     patterns = []
#     start = 0
#     for i in range(1, len(data)):
#         # 값이 다시 0으로 돌아오면 패턴의 끝으로 간주
#         if data[i] == 0 and data[i - 1] != 0:
#             patterns.append((start, i))
#             start = i
#     return patterns


# # 패턴 찾기
# patterns = find_repeating_patterns(a_column_data)

# # 패턴에 따라 데이터를 새 시트에 복사합니다.
# for idx, (start, end) in enumerate(patterns):
#     new_sheet = (
#         f'{"첫번째" if idx == 0 else "두번째" if idx == 1 else "세번째"} 측정데이터'
#     )
#     workbook.create_sheet(new_sheet)
#     new_sheet = workbook[new_sheet]

#     # 원본 시트의 1행과 2행을 복사
#     for row in sheet.iter_rows(min_row=1, max_row=header_row):
#         for cell in row:
#             new_cell = new_sheet.cell(
#                 row=cell.row, column=cell.column, value=cell.value
#             )
#             new_cell.style = cell.style

#     # 패턴에 해당하는 데이터 복사
#     for row in sheet.iter_rows(
#         min_row=start + header_row + 1, max_row=end + header_row + 1
#     ):
#         for cell in row:
#             new_cell = new_sheet.cell(
#                 row=cell.row - start, column=cell.column, value=cell.value
#             )
#             new_cell.style = cell.style

# # 변경 사항 저장
# # workbook.save(file)
# # workbook.close()

# # patterns  # 반복되는 패턴의 시작점과 끝점 인덱스를 반환하여 확인합니다.


# # '첫번째 측정데이터' 시트에서 A열의 데이터를 처리하여,
# # 두 번째 나오는 최대값(24.8)부터 0까지의 데이터를 0부터 최대값 순서로 정렬합니다.

# # 엑셀 파일 로드
# # workbook = load_workbook(file)
# # sheet = workbook["첫번째 측정데이터"]

# # A열 데이터 추출
# a_column_data = [
#     cell.value
#     for cell in sheet["A"]
#     if cell.value is not None and isinstance(cell.value, (int, float))
# ]

# # A열에서 최소값과 최대값 찾기
# min_value = min(a_column_data)
# max_value = max(a_column_data)

# # 최대값이 처음 나타난 위치와 마지막으로 나타난 위치 찾기
# first_max_index = a_column_data.index(max_value) + 1  # 엑셀의 행 번호에 맞추기 위해 +1
# last_max_index = len(a_column_data) - a_column_data[::-1].index(
#     max_value
# )  # 뒤에서부터 찾아서 전체 길이에서 빼기

# # 데이터 재정렬 (두 번째 최대값부터 시작하여 끝까지, 그 다음 시작부터 첫 번째 최대값까지)
# reordered_data = (
#     a_column_data[last_max_index - 1 :] + a_column_data[:first_max_index]
# )

# # 재정렬된 데이터를 시트에 다시 쓰기
# for i, value in enumerate(reordered_data, start=1):
#     sheet.cell(row=i, column=1, value=value)

# # 변경 사항 저장
workbook.save(file)
workbook.close()

# print("첫번째 측정데이터 시트의 A열 데이터가 재정렬되었습니다.")
